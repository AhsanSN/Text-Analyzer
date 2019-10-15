print("started")
import textstat
textstat.set_lang("en_US")
import xlsxwriter
from textblob import TextBlob
from openpyxl import load_workbook

#from xlutils.copy import copy    
#from xlrd import open_workbook
print("imported")

def unique(list1): 
    # intilize a null list 
    unique_list = [] 
      
    # traverse for all elements 
    for x in list1: 
        # check if exists in unique_list or not 
        if x not in unique_list: 
            unique_list.append(x) 
    # print list 
    return unique_list

def charNoSpaces(txt):
    n=0
    for i in txt:
        if(i!= ' '):
            n+=1
    return n

def wordsLongestSentence(txt):
    txt = txt.split(".")
    maax = 0
    for i in txt:
        if(len(i)>maax):
            maax=len(i)
    return(maax)

def wordsShortestSentence(txt):
    txt = txt.split(".")
    maax = 100000
    for i in txt:
        if(len(i)<maax):
            maax=len(i)
    return(maax)

def Average(lst): 
    return sum(lst) / len(lst)

def avgSentenceChars(txt):
    txt = txt.split(".")
    txt_n = []
    for i in (txt):
        txt_n.append(len(i))        
    avg = Average(txt_n)
    return(avg)

def avgSentenceWords(txt):
    txt = txt.split(".")
    txt_n = []
    for i in (txt):
        txt_n.append(len(i.split()))        
    avg = Average(txt_n)
    return(avg)

def avgWordLength(txt):
    txt = txt.split()
    txt_n = []
    for i in (txt):
        txt_n.append(len(i))        
    avg = Average(txt_n)
    return(avg)
    
def syllables(txt):
    n = 0
    n = n + txt.count("a")
    n = n + txt.count("e")
    n = n + txt.count("i")
    n = n + txt.count("o")
    n = n + txt.count("u")
    return n

def words_publisher(txt):
    txt = txt.replace('\n', '')
    return len(txt)/6


workbook_name = 'Stats.xlsx'
wb = load_workbook(workbook_name)
page = wb.active

def main():
    

    txt = str(input("Paste Paragraph: "))
    blob = TextBlob(txt)
    #frames

    #values
    nChars = len(txt)
    nWords = len(txt.split())
    nUniqueWords = len(unique(txt.split()))
    ncharNoSpaces = charNoSpaces(txt)
    nSentences = (txt.count("."))
    wordsLongestSentenceVal = wordsLongestSentence(txt)
    wordsShortestSentenceVal = wordsShortestSentence(txt)
    avgSentenceWordsVal = avgSentenceWords(txt)
    avgSentenceCharsVal = avgSentenceChars(txt)
    avgWordLengthVal = avgWordLength(txt)
    nParagraphsVal = txt.count("\n")
    syllablesVal = syllables(txt)
    words_publisherVal = words_publisher(txt)
    readingLevelVal = textstat.dale_chall_readability_score_v2(txt)
    readingTimeVal = textstat.reading_time(txt)
    sentimentPolarityVal = blob.sentiment.polarity

    nParasInp = int(input("Enter number of paragraphs: "))
    gradeInp = str(input("Enter grade: "))

    '''
    print("charLength", nChars)
    print("nWords", nWords)
    print("nUniqueWords", nUniqueWords)
    print("ncharNoSpaces", ncharNoSpaces)
    print("nSentences", nSentences)
    print("wordsLongestSentence", wordsLongestSentenceVal)
    print("wordsShortestSentence", wordsShortestSentenceVal)
    print("avgSentenceWords", avgSentenceWordsVal)
    print("avgSentenceChars", avgSentenceCharsVal)
    print("avgWordLength", avgWordLengthVal)
    print("nParagraphs", nParagraphsVal)
    print("syllables", syllablesVal)
    print("words_publisher", words_publisherVal)
    print("readingLevel", readingLevelVal)
    print("readingTime", readingTimeVal)
    print("sentimentPolarity", sentimentPolarityVal)
    '''
    print("")
    print("Writing to file")

    # New data to write:
    new_companies = [nChars,nWords,nUniqueWords,ncharNoSpaces ,nSentences , wordsLongestSentenceVal,
                     wordsShortestSentenceVal,avgSentenceWordsVal, avgSentenceCharsVal,  avgWordLengthVal, nParagraphsVal
                     , syllablesVal, words_publisherVal, readingLevelVal, readingTimeVal, sentimentPolarityVal, nParasInp, gradeInp]

    maxR = page.max_row+1
    for column, entry in enumerate(new_companies,start=1):
       page.cell(row=maxR, column=column, value=entry)

    wb.save(filename=workbook_name)

    print("-----------------DONE----------------------")
    main()

   
main()













